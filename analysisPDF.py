#!/usr/bin/env python
# coding: utf-8
import re
import os
import pandas
import shutil
import random as rd
from tqdm import tqdm
import multiprocessing
from tabula import read_pdf
from unidecode import unidecode
from openpyxl import load_workbook 
from concurrent.futures import ProcessPoolExecutor


folder = input('carpeta: ')

pdf_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith('.pdf')]

#traigo mi csv de partidos y de tipos
df_partidos = pandas.read_excel("data/partidos0.xlsx")
df_tipos = pandas.read_excel("data/vehiculos.xlsx")
df_tipos = df_tipos.set_index(['MARCA','MODELO'])

#eliminar espacios y simbolos de la columna PARTIDO

simbolos = ['+', '-', '*', '/', '=', '.', ',',' ']

for simbolo in simbolos:
    df_partidos['LOCALIDAD'] = df_partidos['LOCALIDAD'].str.replace(simbolo, '',regex=True)

#primeros datos extraidos
orden_inicial = ['DOMINIO','MARCA','MODELO','TIPO',"FECHA_TITULARIDAD","NRO_DOCUMENTO","NRO_CUIT","IDENTIFICACION",
                 "CALLE","CP", "LOCALIDAD", "PROVINCIA"] 

#columnas cargadas manualmente
manual = ['EMAIL','CELULAR','TELEFONO']

#datos ya trabajados y separados
orden_final = ['DOMINIO','TIPO_DOCUMENTO','NRO_DOCUMENTO','NRO_CUIT','PROPIETARIO_APELLIDO','PROPIETARIO_NOMBRE',
               'FECHA_TITULARIDAD','CALLE','NUMERO','PISO','DPTO','CP','LOCALIDAD','PROVINCIA','TELEFONO',
               'CELULAR','EMAIL','MARCA','MODELO','TIPO','PARTIDO']

orden = ['MARCA','MODELO','MARCA Y MODELO','TIPO']

#declaro mi dataframe final
df_final = pandas.DataFrame(columns = orden_final)

def pdf_to_excel(path):
    try:
        global df_final, df_tipos
        #analiza el pdf segun un area, se eliminan los valores NaN y se renombra la columna generada
        table = read_pdf(path,
                        pages  = 1,
                        stream = True,
                        area   = (88.872,245.421,811.748,581.573))
        
        df = table[0].dropna()
        df.columns = ['Datos']
        
        #elimina datos innecesarios
        indexNames = [1,2,5,7,8,9,10,12,13,14,15,16]
        df = df.drop(indexNames)
        
        #Convierte la columna de datos en una fila identificado cada elemento por su orden inicial
        df = df.assign(Variables = orden_inicial)
        df = df.set_index('Variables').T

        #Ordena y trabaja los datos iniciales

        for col in ['LOCALIDAD','PROVINCIA']:
            df[col] = df[col].apply(unidecode)
            
            # dni
        doc = df.at['Datos', 'NRO_DOCUMENTO'].split(':')

        df = df.assign(TIPO_DOCUMENTO   = doc[0])
        if int(doc[1]) != 0:
            df.at['Datos', 'NRO_DOCUMENTO'] = doc[1]
        else:
            df.at['Datos', 'NRO_DOCUMENTO'] = ''

            #cuit
        cuit = re.sub(r'[^\w\s]','',df.at['Datos','NRO_CUIT'])
        
        if df.at['Datos','NRO_CUIT'] == '(NO DISPONIBLE)':
            df.at['Datos','NRO_CUIT'] = ''
        else:
            df.at['Datos','NRO_CUIT'] = cuit
        

            #nombre y apellido
        if cuit[0] == '3':
            df = df.assign(PROPIETARIO_APELLIDO = df.at['Datos','IDENTIFICACION'])
            df = df.assign(PROPIETARIO_NOMBRE = '')
        else:
            nombres_list = df.at['Datos','IDENTIFICACION'].split(' ')
            df = df.drop(['IDENTIFICACION'], axis=1)

            df = df.assign(PROPIETARIO_APELLIDO = nombres_list[0])
            nombres_list.pop(0)

            nombre = ' '.join(str(e) for e in nombres_list)
            df = df.assign(PROPIETARIO_NOMBRE = nombre)
            
        if 'NO DISPONIBLE' in df.at['Datos','CP'] or df.at['Datos','CP'] == '9999' or df.at['Datos','CP'] == '0':
            df.at['Datos','CP'] = ''
            
        #     calle numero piso y dpto
        calle_list = df.at['Datos','CALLE'].split(' ')
        
        if 'S/N' not in df.at['Datos','CALLE']:
            try:
                if all(isinstance(int(i), int) for i in calle_list):
                    if len(calle_list) == 3:
                        df = df.assign(CALLE = calle_list[0])
                        df = df.assign(NUMERO = calle_list[1])
                        df = df.assign(PISO = calle_list[2])
                        df = df.assign(DPTO = '')
                    elif len(calle_list) == 2:
                        df = df.assign(CALLE = calle_list[0])
                        df = df.assign(NUMERO = calle_list[1])
                    
            except:
                if calle_list[-1].isdigit():
                    
                    numero = int(calle_list[-1])

                    if  numero > 15:
                        
                        df = df.assign(NUMERO = numero)
                        
                        calle_list.pop(-1)
                        
                        calle = ' '.join(str(e) for e in calle_list)
                        
                        df = df.assign(CALLE = calle)
                        df = df.assign(PISO = '')
                        df = df.assign(DPTO = '')
                        
                    else:
                        try:
                            if calle_list[-2].isdigit():
                                df = df.assign(NUMERO = int(calle_list[-2]))
                                if int(calle_list[-1]) != 0:
                                    df = df.assign(PISO = int(calle_list[-1]))
                                
                                calle_list.pop(-1)
                                calle_list.pop(-1)
                                
                                calle = ' '.join(str(e) for e in calle_list)
                                
                                df = df.assign(CALLE = calle)
                                df = df.assign(DPTO = '')
                            
                        except:
                            df = df.assign(NUMERO = int(calle_list[-1]))
                            
                            calle_list.pop(-1)
                            
                            calle = ' '.join(str(e) for e in calle_list)
                            
                            df = df.assign(CALLE=calle)
                            df = df.assign(PISO = '')
                            df = df.assign(DPTO = '')
                            
                elif calle_list[-1] == 'PB':
                    if calle_list[-2].isdigit():
                        
                        df = df.assign(NUMERO = int(calle_list[-2]))
                        
                        calle_list.pop(-2)
                        
                        calle = ' '.join(str(e) for e in calle_list)
                        
                        df = df.assign(CALLE = calle)
                        df = df.assign(PISO = '')
                        df = df.assign(DPTO = '')

        else:
            numero = rd.randint(100,200)
            df.at['Datos','NUMERO'] = numero
            for i in calle_list:
                if re.search('S/N', i) or i == 'SN':
                    calle_list.remove(i)
                    calle = ' '.join(str(e) for e in calle_list)          

        try:
            if df.at['Datos','PISO'].isdigit():
                if df.at['Datos','PISO'] == '0' or df.at['Datos','PISO'] == '00':
                    df.at['Datos','PISO'] == ''
            else:
                df.at['Datos','PISO'] == ''
        except: 
            df = df.assign(NUMERO = '')
            df = df.assign(PISO = '')
            df = df.assign(DPTO = '')
            
            #tipo
        marca = df.at['Datos','MARCA']
        modelo = df.at['Datos','MODELO']
        
        tipo = df.at['Datos','TIPO']
        
        dominio = df.at['Datos','DOMINIO']
        
        if re.search(r'\w\d\d\d\w\w\w',dominio) or re.search(r'\d\d\d\w\w\w',dominio):
            df.at['Datos','TIPO'] = 4
        elif tipo in ['SEDAN 3 PUERTAS CON PORTON','FURGONETA O UTILITARIO','RURAL 4/5 PUERTAS','RURAL 3 PUERTAS','SEDAN 5 PUERTAS','TODO TERRENO','RURAL 5 PUERTAS','PICK-UP','SEDAN 4 PUERTAS','FURGON', 'PICK-UP CABINA DOBLE','FURGONETA','COUPE','SEDAN 3 PUERTAS','PICK-UP CABINA SIMPLE','DESCAPOTABLE','SEDAN 2 PUERTAS','FAMILIAR','UTILITARIO','PICK-UP CABINA Y MEDIA']:
            df.at['Datos','TIPO'] = 1
        elif tipo in ['PICK-UP CARROZADA','CHASIS C/CABINA DORMITORIO','CHASIS C/CABINA','CHASIS CON CABINA','CAMION HORMIGONERO','CAMION','TRACTOR C/CABINA DORMITORIO','TRACTOR CON CABINA DORMITORIO','CHASIS CON CABINA DORMITORIO','CHASIS CON CABINA DOBLE','TRACTOR DE CARRETERA','OTROS AUTOMOTORES DE CARGA']:
            df.at['Datos','TIPO'] = 2
        elif tipo in ['MIDIBUS','MINIBUS','AUTOBOMBA','AUTOMINIBOMBA','CHASIS S/CABINA','CHASIS SIN CABINA','TRANSPORTE DE PASAJEROS','TRANS.DE PASAJEROS','MINIBUS (O MICROOMNIBUS','MINIBUS','MINIBUS (O MICROOMNIBUS)']:
            df.at['Datos','TIPO'] = 3
        elif tipo == 'MOTOCICLETA':
            df.at['Datos','TIPO'] = 4
        elif tipo in ['REMOLQUE','SEMIRREMOLQUE BITREN D','SEMIRREMOLQUE','SEMIRREMOLQUE BITREN T', 'ACOPLADO','SEMI-ACOPLADO']:
            df.at['Datos','TIPO'] = 5
        elif tipo == 'ESCOLAR':
            df.at['Datos','TIPO'] = 6
        elif tipo == 'COLECIVO':
            df.at['Datos','TIPO'] = 7
        elif tipo in ['SIN ESPECIFICACION','NO INFORMADO','NO TIPIFICADO','AFF']:
            df.at['Datos','TIPO'] = 9
        elif tipo == '-':
            df.at['Datos','TIPO'] = ''
        
        if type(df.at['Datos','TIPO']) == int:
            if not (marca, modelo) in df_tipos.index:
                df_tipos.loc[(marca, modelo), 'TIPO'] = df.at['Datos','TIPO']
        #partido localidad y provincia
        provincia = df.at['Datos','PROVINCIA']
        localidad = df.at['Datos','LOCALIDAD']
        
        df.at['Datos','LOCALIDAD'] = localidad.replace("_"," ")
        df.at['Datos','PROVINCIA'] = provincia.upper()

        provincia = df.at['Datos','PROVINCIA']
        localidad = df.at['Datos','LOCALIDAD']
        
        for simbolo in simbolos:
            localidad = localidad.replace(simbolo,"")
            
        try:
            if provincia == 'CIUDAD AUTONOMA':
                df = df.assign(PARTIDO = 'C.A.B.A.')
                df = df.assign(LOCALIDAD = 'C.A.B.A.')
            else:
                df_localidad = df_partidos[df_partidos['LOCALIDAD'] == localidad]

                df_localidad = df_localidad.drop_duplicates(subset = ['PROVINCIA'])
                df_localidad = df_localidad.set_index('PROVINCIA')

                partido1 = df_localidad.loc[provincia,'PARTIDO']
                df = df.assign(PARTIDO = partido1)
                
        except:
            df = df.assign(PARTIDO = "")
            
        #Ultimo ordenamiento de columnas
        df = df.reindex(columns = df.columns.tolist() + manual)
        df = df[orden_final]

        #Agrega df a df_final
        df_final = pandas.concat([df_final, df])
    except Exception as e:
        #los pdfs que no pudieron analizarse (seguramente problema de area) se envian a revisar
        print(e)
        if not os.path.isdir('revisar_g'):
            os.mkdir('revisar_g')
        shutil.move(path, 'revisar_g')

##Recorre todos los pdf de la carpeta y aplica la funcion a cada uno con tqdm
# num_cores = multiprocessing.cpu_count()

# with ProcessPoolExecutor(max_workers=num_cores) as executor:
#     with tqdm(total=len(pdf_files), desc='Procesando PDFs') as progress:
#         futures = []
        
#         for path in pdf_files:
#             print(path)
#             future = executor.submit(pdf_to_excel, path)
#             future.add_done_callback(lambda future: progress.update())
#             futures.append(future)
        
#         for future in futures:
#             future.result()

#Recorre todos los pdf de la carpeta y aplica la funcion a cada uno
for path in tqdm(pdf_files, desc='Procesando PDFs'):

    try:
        pdf_to_excel(path)

        path = path.split('/')[-1]        

    except Exception as e:
        #los pdfs que no pudieron analizarse (seguramente problema de area) se envian a revisar
        if not os.path.isdir('revisar_g'):
            os.mkdir('revisar_g')
        shutil.move(path, 'revisar_g')

df_final = df_final.replace(';','', regex = True)
df_final = df_final.replace('=','', regex = True)

for col in ['PROPIETARIO_APELLIDO','PROPIETARIO_NOMBRE','CALLE','LOCALIDAD','PROVINCIA','PARTIDO']:
    df_final[col] = df_final[col].apply(unidecode)

df_final = df_final.sort_values(['PARTIDO'], ascending = False)
excel = "dominios.xlsx"

dfd = df_final[['DOMINIO']]

if os.path.isfile(excel):
    #convierte el df en una lista de listas
    datos = df_final.values.tolist()
    dom = dfd.values.tolist()
    #se abre el excel
    wb = load_workbook(excel)
    #indica en que hoja estamos trabajando
    sheets = wb.sheetnames
    
    wsd = wb['dominios']
    for row in dom: 
        wsd.append(row)

    wsm = wb['Sheet1']

    for row in datos: 
        wsm.append(row)
        
    #guarda y cierra el excel
    wb.save(excel)
    
    print("agregados los datos al excel dominios.xlsx")
else:
    with pandas.ExcelWriter(excel) as writer:  

        dfd.to_excel(writer, sheet_name = 'dominios', index = False)

        df_final.to_excel(writer, sheet_name = 'Sheet1', index = False)
        
        print("excel creado y datos agregados")
        
df_tipos.to_excel('data/vehiculos-revisar.xlsx')